import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime
from dataclasses import dataclass
from typing import Dict, List, Set


# AnalysisResult 클래스 정의
@dataclass
class AnalysisResult:
    corner_stats: pd.DataFrame
    ac_stats: pd.DataFrame
    consecutive_stats: pd.DataFrame
    color_stats: pd.DataFrame
    composite_stats: pd.DataFrame
    square_stats: pd.DataFrame
    mirror_stats: pd.DataFrame
    multiple_stats: pd.DataFrame
    prime_stats: pd.DataFrame
    last_digit_stats: pd.DataFrame
    palindrome_stats: pd.DataFrame
    double_stats: pd.DataFrame
    color_comb_stats: pd.DataFrame
    full_report: pd.DataFrame


# LottoExporter 클래스 시작
class LottoExporter:
    """로또 분석 결과 엑셀/CSV 내보내기 클래스"""

    def __init__(self, analysis_result):  # 타입 힌트 제거
        self.result = analysis_result
        self.export_time = datetime.now().strftime("%Y%m%d_%H%M%S")

    def _setup_excel_style(self, ws):
        """엑셀 워크시트 스타일 설정"""
        # 스타일 정의
        header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
        header_font = Font(bold=True, size=11)
        centered = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더 스타일 적용
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = centered
            cell.border = border

        # 데이터 셀 스타일 적용
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = centered
                cell.border = border

        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    def export_to_excel(self, output_dir: str = 'analysis_results'):
        """분석 결과를 엑셀 파일로 내보내기"""
        os.makedirs(output_dir, exist_ok=True)
        wb = Workbook()

        # 기본 통계 시트들
        basic_sheets = [
            ("모서리번호", self.result.corner_stats),
            ("AC값", self.result.ac_stats),
            ("연속패턴", self.result.consecutive_stats),
            ("색상분석", self.result.color_stats),
            ("색상조합", self.result.color_comb_stats),
            ("합성수", self.result.composite_stats),
            ("완전제곱수", self.result.square_stats),
            ("동형수", self.result.mirror_stats),
            ("소수", self.result.prime_stats),
            ("끝수합", self.result.last_digit_stats),
            ("회문수", self.result.palindrome_stats),
            ("쌍수", self.result.double_stats),
        ]

        # 기본 시트들 생성
        for sheet_name, data in basic_sheets:
            if sheet_name == "색상조합":
                # 색상 조합은 특별 처리
                ws = wb.create_sheet(sheet_name)
                data_copy = data.copy()
                self._export_dataframe_to_sheet(ws, data_copy, f"{sheet_name} 분석")
            else:
                ws = wb.create_sheet(sheet_name)
                self._export_dataframe_to_sheet(ws, data, f"{sheet_name} 분석")

        # 배수 분석 특별 처리
        ws = wb.create_sheet("배수")
        multiple_data = self.result.multiple_stats.copy()
        # 컬럼 이름 수정
        multiple_data.columns = ['3배수_횟수', '3배수_비율', '4배수_횟수', '4배수_비율', '5배수_횟수', '5배수_비율']
        self._export_dataframe_to_sheet(ws, multiple_data, "배수 분석")

        # 전체 데이터
        ws = wb.create_sheet("전체데이터")
        self._export_dataframe_to_sheet(ws, self.result.full_report, "전체 분석 데이터")

        # 첫 번째 시트 삭제 (기본 생성되는 빈 시트)
        wb.remove(wb['Sheet'])

        # 파일 저장
        excel_path = f"{output_dir}/로또분석_{self.export_time}.xlsx"
        wb.save(excel_path)
        print(f"엑셀 파일이 저장되었습니다: {excel_path}")

    def export_to_csv(self, output_dir: str = 'analysis_results'):
        """분석 결과를 CSV 파일들로 내보내기"""
        # 결과 저장 디렉토리 생성
        os.makedirs(output_dir, exist_ok=True)

        # 데이터프레임과 파일명 매핑
        exports = {
            'corner_stats': self.result.corner_stats,
            'ac_stats': self.result.ac_stats,
            'consecutive_stats': self.result.consecutive_stats,
            'color_stats': self.result.color_stats,
            'color_comb_stats': self.result.color_comb_stats,
            'composite_stats': self.result.composite_stats,
            'square_stats': self.result.square_stats,
            'mirror_stats': self.result.mirror_stats,
            'multiple_stats': self.result.multiple_stats,
            'prime_stats': self.result.prime_stats,
            'last_digit_stats': self.result.last_digit_stats,
            'palindrome_stats': self.result.palindrome_stats,
            'double_stats': self.result.double_stats,
            'full_report': self.result.full_report
        }

        # CSV 파일들 생성
        for name, df in exports.items():
            csv_path = f"{output_dir}/{name}_{self.export_time}.csv"
            df.to_csv(csv_path, encoding='utf-8-sig')
            print(f"CSV 파일이 저장되었습니다: {csv_path}")

    def _export_dataframe_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """데이터프레임을 워크시트에 저장하고 스타일 적용"""
        # 제목 추가
        ws.append([title])
        ws.merge_cells(f'A1:{get_column_letter(len(df.columns) + 1)}1')
        title_cell = ws['A1']
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal='center')

        # 빈 행 추가
        ws.append([])

        # 데이터프레임 헤더 추가
        headers = [str(col) for col in df.columns]
        if df.index.name:
            headers.insert(0, df.index.name)
        ws.append(headers)

        # 데이터 추가 (값을 문자열로 변환하고 NaN 처리)
        for idx, row in df.iterrows():
            row_data = [idx]
            for val in row:
                try:
                    # Series나 ndarray를 리스트로 변환
                    if isinstance(val, (pd.Series, np.ndarray)):
                        val = val.tolist()

                    # None, NaN 체크
                    if val is None or (isinstance(val, float) and np.isnan(val)):
                        row_data.append('')
                    else:
                        row_data.append(str(val))
                except:
                    # 예외 발생시 빈 문자열 추가
                    row_data.append('')
            ws.append(row_data)

        # 스타일 적용
        self._setup_excel_style(ws)


def update_lotto_analyzer():
    """LottoAnalyzer 클래스에 엑스포트 기능 추가"""

    def export_results(self):
        """분석 결과를 엑셀과 CSV로 내보내기"""
        if not self.analysis_result:
            raise ValueError("먼저 run_full_analysis()를 실행해주세요.")

        exporter = LottoExporter(self.analysis_result)
        exporter.export_to_excel()
        exporter.export_to_csv()

    # LottoAnalyzer 클래스에 메서드 추가
    LottoAnalyzer.export_results = export_results


# 실행 예제
if __name__ == "__main__":
    # LottoAnalyzer 업데이트
    update_lotto_analyzer()

    # 분석기 생성 및 분석 실행
    analyzer = LottoAnalyzer('../lotto.db')
    analyzer.run_full_analysis()

    # 리포트 생성 및 결과 내보내기
    analyzer.generate_report()
    analyzer.export_results()